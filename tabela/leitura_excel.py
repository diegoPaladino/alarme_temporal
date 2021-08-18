# leitura_excel
# fonte: https://medium.com/@andersoneduardo_50576/abrindo-arquivos-excel-em-python-833de325df70

import pandas as pd
import openpyxl
import xlrd


# tentativas com o pandas (sem sucesso, por enquanto):

# df = pd.read_excel("C:/Users/diego/OneDrive/Desktop/teste-leitura-excel-python/Pasta1_teste.xlsx",sheet_name="Planilha1.xls")
# df = xl.parser("Planilha1")
# xl = pd.ExcelFile("C:/Users/diego/OneDrive/Desktop/teste-leitura-excel-python/Pasta1_teste.xls")




# tentativas com o openpyxl:
# tentando ler uma simples planilha de teste báxica (SUCESSO):
# wb = openpyxl.load_workbook(filename="C:/Users/diego/OneDrive/Desktop/teste-leitura-excel-python/Pasta1_teste.xlsx")

# tentando ler a planilha de histórico pronta:
wb = openpyxl.load_workbook(filename="C:/Users/diego/OneDrive/Desktop/teste-leitura-excel-python/JA211-201830050-GERALD_LEVI_DE_LIMA_SANTOS.xlsx")

# wb.sheetnames

# para mostrar a planilha (simples - SUCESSO):
# for d in wb['Planilha1'].iter_rows(values_only = True):
        # print(d)

# para mostrar o histórico (UNSUCDESSFUL)
for d in wb['1-HISTÓRICO (2)'].iter_rows(values_only = True):
        print(d)


# ####################
# tentativa com o xlrd:

# wb = xlrd.open_workbook("C:/Users/diego/OneDrive/Desktop/teste-leitura-excel-python/Pasta1_teste.xlsx")

# for i in range(wb.sheet_by_name('Planilha1').nrows):
#     print(wb.sheet_name('Planilha1').row_values(i))